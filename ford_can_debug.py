#!/usr/bin/env python3
"""
ford_can_debug.py — decode Ford CAN traffic with opendbc, save all raw data,
highlight what changed, and export "events" (door open, lock, seatbelt, gear...)
to a separate sheet with cansend commands so you can resend them to the car.

Two input modes:
  * CSV file      — a previously captured dump.
  * LIVE port(s)  — read straight off one or more CAN interfaces in real time
                    (SocketCAN, Linux). Multiple interfaces are read together.

Output workbook (3 sheets):
  * Raw_Log  — EVERY frame: time/bus/ID/bytes/decoded signals. Changed bytes are
               highlighted yellow; event frames are highlighted orange.
  * Changes  — only the frames where a payload byte changed vs the prior same-ID
               frame (the reverse-engineering view).
  * Events   — discrete state transitions (Driver door OPENED, Seatbelt BUCKLED,
               Gear -> R, ...) each with the exact `cansend` line to replay it.
A plain-text raw dump (all frames) is also written next to the workbook.

Decoding is MANDATORY (uses opendbc's Ford database).

Install:
    pip install opendbc openpyxl
    # can-utils gives you cansend/candump:  sudo apt install can-utils

Bring a live interface up first (example, 500 kbit/s HS-CAN):
    sudo ip link set can0 up type can bitrate 500000
    sudo ip link set can1 up type can bitrate 500000

Examples:
    # From a CSV dump:
    python ford_can_debug.py --csv dump.csv

    # Live from one port until you press Ctrl-C:
    python ford_can_debug.py --live can0

    # Live from TWO ports at once, for 30 seconds:
    python ford_can_debug.py --live can0 can1 --duration 30

    # Live, stop after 5000 frames, custom output name:
    python ford_can_debug.py --live can0 --max-frames 5000 --out door_test.xlsx
"""

from __future__ import annotations
import argparse
import csv
import os
import select
import socket
import struct
import sys
import time

# --------------------------------------------------------------------------- #
# Ford defaults
# --------------------------------------------------------------------------- #

DEFAULT_DBC = "ford_lincoln_base_pt"   # the Ford/Lincoln database shipped in opendbc

# Known-good message.signal pairs opendbc itself reads for Ford body/comfort state.
# {message_name: {signal_name: (friendly_label, {value: meaning} or None)}}
FORD_EVENT_SIGNALS = {
    "BodyInfo_3_FD1": {
        "DrStatDrv_B_Actl":   ("Driver door",     {0: "CLOSED", 1: "OPEN"}),
        "DrStatPsngr_B_Actl": ("Passenger door",  {0: "CLOSED", 1: "OPEN"}),
        "DrStatRl_B_Actl":    ("Rear-left door",  {0: "CLOSED", 1: "OPEN"}),
        "DrStatRr_B_Actl":    ("Rear-right door", {0: "CLOSED", 1: "OPEN"}),
    },
    "RCMStatusMessage2_FD1": {
        "FirstRowBuckleDriver": ("Driver seatbelt", {0: "UNBUCKLED", 1: "BUCKLED"}),
    },
    "PowertrainData_10": {
        "TrnRng_D_Rq": ("Gear selector", None),
    },
    "BCM_Lamp_Stat_FD1": {
        "RvrseLghtOn_B_Stat": ("Reverse light", {0: "OFF", 1: "ON"}),
    },
}

# Any decoded signal whose name contains one of these is ALSO auto-watched as an
# event (covers locks/windows/ignition/trunk defined in the DBC that the curated
# list above doesn't name explicitly).
EVENT_KEYWORDS = (
    "door", "ajar", "lock", "unlock", "latch",
    "belt", "buckle", "ignition", "ignsts", "ignition_status",
    "gear", "trnrng", "window", "wndw", "reverse", "rvrse",
    "trunk", "liftgate", "decklid", "hood", "keyfob", "remote", "rke",
)


# --------------------------------------------------------------------------- #
# Frame
# --------------------------------------------------------------------------- #

class Frame:
    __slots__ = ("t", "addr", "bus", "data")
    def __init__(self, t, addr, bus, data):
        self.t = t; self.addr = addr; self.bus = bus; self.data = data


# --------------------------------------------------------------------------- #
# CSV input — flexible, auto-detects common dump layouts
# --------------------------------------------------------------------------- #

def _to_int(x: str) -> int:
    x = x.strip()
    if not x:
        return 0
    if x.lower().startswith("0x"):
        return int(x, 16)
    try:
        return int(x, 16)
    except ValueError:
        return int(x, 10)


def _parse_data(cells):
    if len(cells) == 1:
        s = cells[0].strip().replace("0x", "").replace(",", " ").replace("-", " ")
        if " " in s:
            return bytes(int(b, 16) for b in s.split() if b)
        if s and len(s) % 2 == 0:
            return bytes.fromhex(s)
        return b""
    out = bytearray()
    for c in cells:
        c = c.strip().replace("0x", "")
        if c:
            out.append(int(c, 16) & 0xFF)
    return bytes(out)


def read_csv(path):
    """Yield Frame objects from a CSV dump (auto-detects columns)."""
    with open(path, newline="") as f:
        sample = f.read(4096); f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(f, dialect))
    if not rows:
        return

    header = [h.strip().lower() for h in rows[0]]
    has_header = any(any(c.isalpha() for c in h) and "0x" not in h for h in header)

    def find(names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    if has_header:
        ti = find({"time", "timestamp", "time stamp", "ts", "secs", "seconds"})
        idi = find({"id", "arbitration id", "arbid", "arb id", "canid", "can id", "pgn"})
        busi = find({"bus", "channel", "chan", "ch"})
        single = find({"data", "payload", "bytes", "data bytes"})
        data_cols = [single] if single is not None else [
            i for i, h in enumerate(header)
            if (h.startswith(("d", "b")) and h[1:].isdigit()) or h.startswith("byte")
        ]
        start = 1
    else:
        ti, idi, busi = 0, 1, None
        data_cols = list(range(2, len(rows[0])))
        start = 0

    if idi is None:
        idi = 1 if len(header) > 1 else 0
    if not data_cols:
        data_cols = [len(header) - 1]

    for r in rows[start:]:
        if not r or len(r) <= idi:
            continue
        try:
            t = float(r[ti]) if ti is not None and r[ti].strip() else 0.0
        except ValueError:
            t = 0.0
        try:
            addr = _to_int(r[idi])
        except ValueError:
            continue
        bus = _to_int(r[busi]) if busi is not None and busi < len(r) else 0
        data = _parse_data([r[i] for i in data_cols if i < len(r)])
        yield Frame(t, addr, bus, data)


# --------------------------------------------------------------------------- #
# LIVE input — raw SocketCAN, multiple interfaces via one select() loop
# --------------------------------------------------------------------------- #

# struct can_frame { canid_t can_id; u8 can_dlc; u8 __pad; u8 __res0; u8 len8_dlc; u8 data[8]; }
CAN_FRAME_SIZE = 16
CAN_EFF_FLAG = 0x80000000     # extended (29-bit) id
CAN_RTR_FLAG = 0x40000000     # remote request
CAN_ERR_FLAG = 0x20000000     # error frame
CAN_EFF_MASK = 0x1FFFFFFF
CAN_SFF_MASK = 0x000007FF


def open_can_socket(channel):
    """Open and bind a raw SocketCAN socket on the given interface (e.g. 'can0')."""
    if not hasattr(socket, "AF_CAN"):
        sys.exit("SocketCAN is Linux-only; live capture is unavailable on this OS.")
    s = socket.socket(socket.AF_CAN, socket.SOCK_RAW, socket.CAN_RAW)
    try:
        s.bind((channel,))
    except OSError as e:
        sys.exit(f"Cannot bind to '{channel}': {e}. Is the interface up? "
                 f"Try:  sudo ip link set {channel} up type can bitrate 500000")
    return s


def capture_live(channels, duration, max_frames):
    """
    Yield Frame objects read live from one or more CAN interfaces.
    bus number = position of the interface in `channels` (can0->0, can1->1, ...).
    Stops on Ctrl-C, after `duration` seconds, or after `max_frames` frames.
    """
    socks = [open_can_socket(ch) for ch in channels]
    bus_of = {s.fileno(): i for i, s in enumerate(socks)}
    start = time.time()
    count = 0
    print(f"Listening on {', '.join(channels)} — Ctrl-C to stop.", file=sys.stderr)
    try:
        while True:
            if duration and (time.time() - start) >= duration:
                break
            if max_frames and count >= max_frames:
                break
            ready, _, _ = select.select(socks, [], [], 0.2)
            for s in ready:
                raw = s.recv(CAN_FRAME_SIZE)
                if len(raw) < CAN_FRAME_SIZE:
                    continue
                can_id, dlc = struct.unpack_from("<IB", raw, 0)
                if can_id & (CAN_ERR_FLAG | CAN_RTR_FLAG):
                    continue                      # skip error/remote frames
                addr = can_id & (CAN_EFF_MASK if can_id & CAN_EFF_FLAG else CAN_SFF_MASK)
                data = raw[8:8 + (dlc & 0x0F)]
                yield Frame(time.time() - start, addr, bus_of[s.fileno()], bytes(data))
                count += 1
    finally:
        for s in socks:
            s.close()


# --------------------------------------------------------------------------- #
# opendbc decode (mandatory) — one parser per bus, built lazily
# --------------------------------------------------------------------------- #

def dbc_messages(dbc_name):
    """Return ([(addr, name), ...], {addr: name}) by reading the shipped .dbc."""
    try:
        import opendbc
    except ImportError:
        sys.exit("opendbc is required. Install with:  pip install opendbc")
    dbc_dir = os.path.join(os.path.dirname(opendbc.__file__), "dbc")
    path = os.path.join(dbc_dir, dbc_name + ".dbc")
    if not os.path.exists(path):
        avail = sorted(x[:-4] for x in os.listdir(dbc_dir)
                       if x.endswith(".dbc") and "ford" in x.lower())
        sys.exit(f"DBC not found: {path}\nFord DBCs available: {avail}")
    pairs, addr2name = [], {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line.startswith("BO_ "):
                p = line.split()
                if len(p) >= 3:
                    addr = int(p[1]) & CAN_EFF_MASK
                    name = p[2].rstrip(":")
                    pairs.append((addr, name)); addr2name[addr] = name
    return pairs, addr2name


class ParserPool:
    """Holds one opendbc CANParser per bus so every captured bus decodes correctly."""
    def __init__(self, dbc_name, pairs):
        try:
            from opendbc.can.parser import CANParser
        except ImportError:
            sys.exit("opendbc is required. Install with:  pip install opendbc")
        self._CANParser = CANParser
        self.dbc = dbc_name
        self.messages = [(name, 0) for _, name in pairs]   # freq 0 = don't enforce rate
        self.cache = {}

    def _get(self, bus):
        if bus not in self.cache:
            self.cache[bus] = self._CANParser(self.dbc, self.messages, bus)
        return self.cache[bus]

    def decode(self, f):
        """Decode one frame -> {signal: value}."""
        p = self._get(f.bus)
        p.update([(int(f.t * 1e9), [(f.addr, f.data, f.bus)])])
        try:
            return dict(p.vl[f.addr])
        except (KeyError, TypeError):
            return {}


# --------------------------------------------------------------------------- #
# cansend formatting + event helpers
# --------------------------------------------------------------------------- #

def cansend(channel, addr, data):
    id_str = f"{addr:03X}" if addr <= 0x7FF else f"{addr:08X}"
    return f"cansend {channel} {id_str}#{data.hex().upper()}"


def panda_cmd(bus, addr, data):
    """comma-panda equivalent of a cansend line (no SocketCAN needed)."""
    return f'p.can_send(0x{addr:X}, bytes.fromhex("{data.hex().upper()}"), {bus})'


def label_value(val, mapping):
    r = round(val, 3)
    iv = int(r) if float(r).is_integer() else r
    if mapping and iv in mapping:
        return f"{mapping[iv]} ({iv})"
    return str(iv)


def is_watched(msg_name, sig_name):
    """Return (label, mapping) if this signal is an event signal, else None."""
    m = FORD_EVENT_SIGNALS.get(msg_name)
    if m and sig_name in m:
        return m[sig_name]
    low = sig_name.lower()
    if any(k in low for k in EVENT_KEYWORDS):
        return (sig_name, None)
    return None


# --------------------------------------------------------------------------- #
# Processor — feeds frames one at a time; builds workbook + raw dump; live prints
# --------------------------------------------------------------------------- #

class Processor:
    def __init__(self, pool, addr2name, channel_for, out_xlsx, out_raw, live=False):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            sys.exit("openpyxl is required. Install with:  pip install openpyxl")

        self.pool = pool
        self.addr2name = addr2name
        self.channel_for = channel_for          # bus -> interface name for cansend
        self.out_xlsx = out_xlsx
        self.live = live

        self.YELLOW = PatternFill("solid", fgColor="FFF2A6")
        self.ORANGE = PatternFill("solid", fgColor="FFC46B")
        self.GREEN  = PatternFill("solid", fgColor="B7E1A1")
        self.HEADER = PatternFill("solid", fgColor="305496")
        self.HFONT  = Font(color="FFFFFF", bold=True)
        self._Alignment = Alignment

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Raw_Log"
        self.raw_hdr = ["idx", "t(s)", "bus", "ID", "DLC"] + [f"B{i}" for i in range(8)] \
                       + ["Decoded", "Event"]
        self.ws.append(self.raw_hdr)

        self.ws_ch = self.wb.create_sheet("Changes")
        self.ws_ch.append(["idx", "t(s)", "bus", "ID", "Changed bytes", "Prev payload",
                           "New payload", "Decoded", "cansend (resend)", "panda (p.can_send)"])

        self.ws_ev = self.wb.create_sheet("Events")
        self.ws_ev.append(["t(s)", "bus", "Event", "ID", "Message", "From", "To",
                           "Payload", "cansend (resend to car)", "panda (p.can_send)"])

        self.raw_fh = open(out_raw, "w")
        self.out_raw = out_raw

        self.prev_payload = {}   # (bus, addr) -> last payload
        self.last_sig = {}       # (bus, addr, sig) -> last value
        self.idx = 0
        self.n_changes = 0
        self.n_events = 0

    def process(self, f):
        sigs = self.pool.decode(f)
        msg_name = self.addr2name.get(f.addr, "")
        dec_str = "  ".join(f"{k}={v:g}" for k, v in sigs.items())
        chan = self.channel_for(f.bus)

        # ---- event detection ----
        frame_events = []
        for sig, val in sigs.items():
            w = is_watched(msg_name, sig)
            if not w:
                continue
            key = (f.bus, f.addr, sig)
            prev = self.last_sig.get(key)
            cur = round(val, 3)
            if prev is not None and prev != cur:
                lbl, mp = w
                frame_events.append((lbl, msg_name,
                                     label_value(prev, mp), label_value(cur, mp)))
            self.last_sig[key] = cur

        # ---- byte-change detection ----
        pv = self.prev_payload.get((f.bus, f.addr))
        changed = [i for i in range(min(len(f.data), len(pv or b"")))
                   if f.data[i] != pv[i]] if pv is not None else []

        # ---- Raw_Log row ----
        row = [self.idx, round(f.t, 4), f.bus, f"0x{f.addr:X}", len(f.data)]
        for i in range(8):
            row.append(f"{f.data[i]:02X}" if i < len(f.data) else "")
        ev_text = "; ".join(f"{lbl}: {a}->{b}" for lbl, _, a, b in frame_events)
        row += [dec_str, ev_text]
        self.ws.append(row)
        r = self.ws.max_row
        for i in changed:
            self.ws.cell(r, 6 + i).fill = self.YELLOW
        if frame_events:
            self.ws.cell(r, 4).fill = self.ORANGE
            self.ws.cell(r, len(self.raw_hdr)).fill = self.ORANGE

        # ---- raw text dump (streamed) ----
        self.raw_fh.write(
            f"{f.t:>12.4f}  bus{f.bus}  0x{f.addr:03X}#{f.data.hex().upper():<16}"
            + (f"  | {dec_str}" if dec_str else "")
            + (f"  ** {ev_text}" if ev_text else "") + "\n"
        )

        # ---- Changes sheet ----
        if changed:
            self.n_changes += 1
            self.ws_ch.append([self.idx, round(f.t, 4), f.bus, f"0x{f.addr:X}",
                               ",".join(map(str, changed)),
                               (pv or b"").hex().upper(), f.data.hex().upper(),
                               dec_str, cansend(chan, f.addr, f.data),
                               panda_cmd(f.bus, f.addr, f.data)])

        # ---- Events sheet (+ live console) ----
        for lbl, mname, a, b in frame_events:
            self.n_events += 1
            cmd = cansend(chan, f.addr, f.data)
            pcmd = panda_cmd(f.bus, f.addr, f.data)
            self.ws_ev.append([round(f.t, 4), f.bus, lbl, f"0x{f.addr:X}", mname,
                               a, b, f.data.hex().upper(), cmd, pcmd])
            self.ws_ev.cell(self.ws_ev.max_row, 3).fill = self.GREEN
            if self.live:
                print(f"[{f.t:7.2f}s] {chan}  {lbl}: {a} -> {b}")
                print(f"            cansend: {cmd}")
                print(f"            panda  : {pcmd}")

        self.prev_payload[(f.bus, f.addr)] = f.data
        self.idx += 1

    def finalize(self):
        for sheet in (self.ws, self.ws_ch, self.ws_ev):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(1, c)
                cell.fill = self.HEADER; cell.font = self.HFONT
                cell.alignment = self._Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
        self.ws.column_dimensions["N"].width = 60
        self.ws.column_dimensions["O"].width = 34
        self.ws_ch.column_dimensions["I"].width = 34
        self.ws_ch.column_dimensions["J"].width = 46
        self.ws_ev.column_dimensions["C"].width = 20
        self.ws_ev.column_dimensions["H"].width = 20
        self.ws_ev.column_dimensions["I"].width = 34
        self.ws_ev.column_dimensions["J"].width = 46

        self.wb.save(self.out_xlsx)
        self.raw_fh.close()

        print(f"\nFrames:        {self.idx}")
        print(f"Byte-changes:  {self.n_changes}  (sheet 'Changes')")
        print(f"Events:        {self.n_events}  (sheet 'Events', with cansend commands)")
        print(f"Workbook:      {self.out_xlsx}")
        print(f"Raw dump:      {self.out_raw}")


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser(
        description="Decode Ford CAN traffic (CSV or live port) and export raw/changes/events to Excel.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv", help="CSV file of CAN frames")
    src.add_argument("--live", nargs="+", metavar="IFACE",
                     help="live capture from one or more SocketCAN interfaces (e.g. --live can0 can1)")
    ap.add_argument("--dbc", default=DEFAULT_DBC, help=f"opendbc DBC name (default {DEFAULT_DBC})")
    ap.add_argument("--duration", type=float, default=0, help="live: stop after N seconds")
    ap.add_argument("--max-frames", type=int, default=0, help="live: stop after N frames")
    ap.add_argument("--channel", default="can0",
                    help="CSV mode: interface name to put in cansend commands (default can0)")
    ap.add_argument("--out", default="ford_capture.xlsx", help="output .xlsx path")
    ap.add_argument("--raw", default=None, help="raw text dump path (default: <out>.raw.txt)")
    args = ap.parse_args()

    pairs, addr2name = dbc_messages(args.dbc)
    pool = ParserPool(args.dbc, pairs)
    out_raw = args.raw or (os.path.splitext(args.out)[0] + ".raw.txt")

    if args.live:
        # bus i -> interface name = the i-th interface given on the command line.
        channel_for = lambda b: args.live[b] if 0 <= b < len(args.live) else f"can{b}"
        source = capture_live(args.live, args.duration, args.max_frames)
        live = True
    else:
        channel_for = lambda b: args.channel
        source = read_csv(args.csv)
        live = False

    print(f"Loaded '{args.dbc}' ({len(pairs)} messages).")
    proc = Processor(pool, addr2name, channel_for, args.out, out_raw, live=live)
    try:
        for f in source:
            proc.process(f)
    except KeyboardInterrupt:
        print("\nStopped by user.", file=sys.stderr)
    finally:
        proc.finalize()


if __name__ == "__main__":
    main()
